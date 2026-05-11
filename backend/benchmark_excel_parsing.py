#!/usr/bin/env python3
"""
⚡ Script de benchmark : Comparer les performances de parsing Excel
Utilisation: python3 benchmark_excel_parsing.py <fichier.xlsx>
"""

import sys
import time
import io
import openpyxl
from pathlib import Path

# Import optionnel
try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def benchmark_openpyxl_original(filepath, sheet_name):
    """Version ORIGINALE: double boucle sur TOUTES les cellules"""
    start = time.time()
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb[sheet_name]
    
    count = 0
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if sheet.cell(row=r, column=c).value is not None:
                count += 1
    
    wb.close()
    elapsed = time.time() - start
    return elapsed, count


def benchmark_openpyxl_optimized(filepath, sheet_name):
    """Version OPTIMISÉE: limiter aux colonnes avec contenu"""
    start = time.time()
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb[sheet_name]
    
    # Déterminer les colonnes/lignes avec contenu
    occupied_columns = set()
    occupied_rows = set()
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                occupied_columns.add(cell.column)
                occupied_rows.add(cell.row)
    
    if occupied_columns and occupied_rows:
        min_col = min(occupied_columns)
        max_col = max(occupied_columns)
        min_row = min(occupied_rows)
        max_row = max(occupied_rows)
        
        count = 0
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if sheet.cell(row=r, column=c).value is not None:
                    count += 1
    else:
        count = 0
    
    wb.close()
    elapsed = time.time() - start
    return elapsed, count


def benchmark_pandas(filepath, sheet_name):
    """Version pandas: read_excel simple"""
    start = time.time()
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    count = df.size
    elapsed = time.time() - start
    return elapsed, count


def benchmark_polars(filepath, sheet_name):
    """Version POLARS: ultra-rapide!"""
    start = time.time()
    df = pl.read_excel(filepath, sheet_name=sheet_name)
    count = df.shape[0] * df.shape[1]
    elapsed = time.time() - start
    return elapsed, count


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 benchmark_excel_parsing.py <fichier.xlsx> [sheet_name]")
        print("\nExemple: python3 benchmark_excel_parsing.py data.xlsx Sheet1")
        sys.exit(1)
    
    filepath = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else "Sheet1"
    
    if not Path(filepath).exists():
        print(f"❌ Fichier non trouvé: {filepath}")
        sys.exit(1)
    
    file_size_mb = Path(filepath).stat().st_size / (1024 * 1024)
    print(f"\n📊 Benchmark de parsing Excel")
    print(f"📁 Fichier: {filepath}")
    print(f"📏 Taille: {file_size_mb:.2f} MB")
    print(f"📋 Feuille: {sheet_name}")
    print("=" * 60)
    
    # Test openpyxl original
    print("\n⏱️  openpyxl (ORIGINAL - double boucle)...")
    try:
        elapsed, count = benchmark_openpyxl_original(filepath, sheet_name)
        print(f"   ✅ {elapsed:.3f}s pour parcourir {count} cellules")
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
    
    # Test openpyxl optimisé
    print("\n⏱️  openpyxl (OPTIMISÉ - colonnes limitées)...")
    try:
        elapsed, count = benchmark_openpyxl_optimized(filepath, sheet_name)
        speedup_vs_original = benchmark_openpyxl_original(filepath, sheet_name)[0] / elapsed
        print(f"   ✅ {elapsed:.3f}s (⚡ {speedup_vs_original:.1f}x plus rapide)")
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
    
    # Test pandas
    if PANDAS_AVAILABLE:
        print("\n⏱️  pandas.read_excel()...")
        try:
            elapsed, count = benchmark_pandas(filepath, sheet_name)
            print(f"   ✅ {elapsed:.3f}s")
        except Exception as e:
            print(f"   ❌ Erreur: {e}")
    else:
        print("\n⏱️  pandas: ⚠️  Non installé")
    
    # Test polars
    if POLARS_AVAILABLE:
        print("\n⏱️  polars.read_excel() ⚡⚡⚡...")
        try:
            elapsed, count = benchmark_polars(filepath, sheet_name)
            speedup_vs_openpyxl = benchmark_openpyxl_original(filepath, sheet_name)[0] / elapsed
            print(f"   ✅ {elapsed:.3f}s (⚡⚡⚡ {speedup_vs_openpyxl:.0f}x plus rapide que openpyxl!)")
        except Exception as e:
            print(f"   ❌ Erreur: {e}")
    else:
        print("\n⏱️  polars: ⚠️  Non installé (installez avec: pip install polars pyarrow)")
    
    print("\n" + "=" * 60)
    print("💡 Recommandation: Installer polars pour 50-100x plus rapide!")
    print("   pip install polars pyarrow")


if __name__ == "__main__":
    main()
