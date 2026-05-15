# backend/xlsx_parser.py
#
# Architecture refactorisée selon le schéma proposé :
#   Lecture  → Polars + Calamine (binaire ultra-rapide)
#   Stockage → Parquet temporaire (côté backend)
#   Indexation → ChromaDB sur métadonnées de colonnes uniquement (pas les lignes)
#   Dialogue → FastAPI + code-gen (le LLM génère du code Polars exécuté côté backend)

import os
import uuid
import tempfile
import openpyxl
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import polars as pl
from langchain_core.documents import Document
import chromadb
from langchain_community.vectorstores import Chroma
from langchain_community.embeddings import OllamaEmbeddings

URL_OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434")

# Répertoire temp pour les fichiers Parquet (nettoyé à chaque nouvelle session)
PARQUET_TEMP_DIR = Path(tempfile.gettempdir()) / "excel_rag_parquet"
PARQUET_TEMP_DIR.mkdir(exist_ok=True)

# ── Clients ChromaDB éphémères par session ──────────────────────────────────
_csv_clients: Dict[str, chromadb.EphemeralClient] = {}


# ---------------------------------------------------------------------------
# Utilitaires cellules fusionnées (openpyxl, fallback uniquement)
# ---------------------------------------------------------------------------

def get_merged_context(sheet, row, col):
    """Vérifie si une cellule est fusionnée et renvoie sa valeur et sa plage."""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return val, merged_range
    return None, None


def extract_table_with_strict_bounds(sheet, start_row, start_col, visited):
    """Extrait un bloc de données en isolant le titre fusionné (openpyxl fallback)."""
    title, m_range = get_merged_context(sheet, start_row, start_col)
    actual_start_row = m_range.max_row + 1 if m_range else start_row

    if m_range:
        for r in range(m_range.min_row, m_range.max_row + 1):
            for c in range(m_range.min_col, m_range.max_col + 1):
                visited.add((r, c))

    max_c = start_col
    while sheet.cell(row=actual_start_row, column=max_c).value is not None:
        max_c += 1

    if max_c == start_col:
        return None, None, visited

    table_data = []
    curr_row = actual_start_row

    while curr_row <= sheet.max_row:
        row_values = []
        is_row_empty = True

        test_val, test_merge = get_merged_context(sheet, curr_row, start_col)
        if curr_row > actual_start_row and test_merge:
            break

        for c in range(start_col, max_c):
            val = sheet.cell(row=curr_row, column=c).value
            if val is not None:
                is_row_empty = False
            row_values.append(val)

        if is_row_empty:
            break

        table_data.append(row_values)
        for c in range(start_col, max_c):
            visited.add((curr_row, c))
        curr_row += 1

    return table_data, title, visited


# ---------------------------------------------------------------------------
# Détection des blocs (Polars)
# ---------------------------------------------------------------------------

def obtenir_masque_densite(df: pl.DataFrame):
    # Crée un DataFrame de booléens : True si la cellule n'est pas nulle
    return df.select(pl.all().is_not_null())


def identifier_ilots_donnees(df: pl.DataFrame) -> list[dict]:
    """
    Identifie les coordonnées (lignes/colonnes) des blocs de données.
    ✅ Considère aussi les COLONNES vides comme des séparateurs (tableaux alignés horizontalement).
    """
    if df.is_empty():
        return []

    # 1. Création du masque de présence
    masque = df.select([
        pl.all().is_not_null() & (pl.all().cast(pl.Utf8) != "")
    ])

    # 2. Analyse des lignes pleines (au moins une valeur non-null)
    lignes_pleines = masque.select(pl.any_horizontal(pl.all())).to_series().to_list()
    
    blocs = []
    start_row = None
    
    for i, est_pleine in enumerate(lignes_pleines):
        if est_pleine and start_row is None:
            start_row = i
        elif not est_pleine and start_row is not None:
            # ✅ Fin d'un bloc vertical
            df_temp = masque.slice(start_row, i - start_row)
            
            # Trouver les colonnes qui ont des données dans CE bloc
            cols_with_data = [
                j for j in range(len(df_temp.columns))
                if df_temp[df_temp.columns[j]].any()
            ]
            
            if cols_with_data:
                # ✅ Créer des sous-blocs en groupant les colonnes contiguës
                # (séparation par colonnes vides)
                start_col = cols_with_data[0]
                
                for idx in range(len(cols_with_data)):
                    curr_col = cols_with_data[idx]
                    next_col = cols_with_data[idx + 1] if idx + 1 < len(cols_with_data) else None
                    
                    # Vérifier s'il y a un trou (colonne vide) après curr_col
                    if next_col is not None and next_col - curr_col > 1:
                        # Trou = fin d'un bloc horizontal
                        end_col = curr_col + 1
                        blocs.append({
                            "start_row": start_row,
                            "end_row": i,
                            "start_col": start_col,
                            "end_col": end_col
                        })
                        start_col = next_col
                
                # Ajouter le dernier sous-bloc
                if start_col is not None:
                    end_col = cols_with_data[-1] + 1
                    blocs.append({
                        "start_row": start_row,
                        "end_row": i,
                        "start_col": start_col,
                        "end_col": end_col
                    })
            
            start_row = None
    
    # Gestion du dernier bloc
    if start_row is not None:
        df_temp = masque.slice(start_row, len(masque) - start_row)
        cols_with_data = [
            j for j in range(len(df_temp.columns))
            if df_temp[df_temp.columns[j]].any()
        ]
        
        if cols_with_data:
            # Même logique
            start_col = cols_with_data[0]
            
            for idx in range(len(cols_with_data)):
                curr_col = cols_with_data[idx]
                next_col = cols_with_data[idx + 1] if idx + 1 < len(cols_with_data) else None
                
                if next_col is not None and next_col - curr_col > 1:
                    end_col = curr_col + 1
                    blocs.append({
                        "start_row": start_row,
                        "end_row": len(masque),
                        "start_col": start_col,
                        "end_col": end_col
                    })
                    start_col = next_col
            
            if start_col is not None:
                end_col = cols_with_data[-1] + 1
                blocs.append({
                    "start_row": start_row,
                    "end_row": len(masque),
                    "start_col": start_col,
                    "end_col": end_col
                })

    return blocs


#def detecter_blocs_continus(df: pl.DataFrame) -> int:
#    """Compte le nombre de blocs de données séparés par des lignes vides (Polars)."""
#    if df.is_empty():
#        return 0
#    # Une ligne est "vide" si toutes ses valeurs sont nulles
#    null_mask = pl.Series([
#        all(v is None for v in row)
#        for row in df.iter_rows()
#    ])
#    # Compter les transitions False→True pour dénombrer les blocs
#    blocs = 0
#    in_bloc = False
#    for is_null in null_mask:
#        if not is_null and not in_bloc:
#            blocs += 1
#            in_bloc = True
#        elif is_null:
#            in_bloc = False
#    return blocs

# ---------------------------------------------------------------------------
# Lecture principale : Polars + Calamine (rapide) → fallback openpyxl
# ---------------------------------------------------------------------------

def _read_excel_with_original_row_numbers(uploaded_file, sheet_name: str) -> Tuple[pl.DataFrame, List[int]]:
    """
    Lit un fichier Excel et retourne les VRAIS numéros de ligne Excel (1-basés, incluant les lignes vides).
    
    Retourne (df_polars, original_row_numbers)
    où original_row_numbers[i] = le numéro de ligne Excel réel pour la ligne i du DataFrame
    """
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[sheet_name]
    
    # Créer un mapping : row_excel -> data
    original_row_numbers = []
    rows_data = []
    
    for excel_row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        original_row_numbers.append(excel_row_idx)
        rows_data.append(row)
    
    # Convertir en Polars
    # Déterminer le nombre de colonnes
    max_cols = max(len(row) for row in rows_data) if rows_data else 0
    
    # Normaliser toutes les lignes au même nombre de colonnes
    # ✅ Convertir les valeurs en string pour éviter les problèmes de schéma mixte
    normalized_rows = []
    for row in rows_data:
        # Convertir chaque valeur en string pour éviter les incohérences de type
        # (datetime vs string, float vs int, etc.)
        normalized_row = [
            str(v) if v is not None else None 
            for v in row
        ] + [None] * (max_cols - len(row))
        normalized_rows.append(normalized_row)
    
    # Créer un DataFrame Polars avec inférence de schéma sur tout le dataset
    # infer_schema_length=None = inférer sur TOUTES les lignes (robuste pour les gros fichiers)
    try:
        df = pl.DataFrame(
            normalized_rows,
            schema=[f"col_{i}" for i in range(max_cols)] if max_cols > 0 else [],
            infer_schema_length=None  # ✅ Inférer sur TOUT le dataset, pas juste les 100 premières
        )
    except Exception as e:
        print(f"⚠️  Fallback: création du DataFrame avec tous les types en string")
        # Si ça échoue encore, forcer tous les types en string
        df = pl.DataFrame(
            normalized_rows,
            schema={f"col_{i}": pl.Utf8 for i in range(max_cols)} if max_cols > 0 else {}
        )
    
    return df, original_row_numbers


def find_tables_in_sheet(uploaded_file, sheet_name: str):
    """
    Version Segmentée : Découpe la feuille en unités logiques.
    
    ✅ IMPORTANT: Préserve les numéros de ligne Excel ORIGINAUX (avec les lignes vides)
    avant tout traitement de compaction. Cela permet une segmentation correcte.
    """
    try:
        # ✅ Lire avec openpyxl pour obtenir les VRAIS numéros de ligne Excel
        df_raw, original_row_numbers = _read_excel_with_original_row_numbers(uploaded_file, sheet_name)
        
        # Ajouter une colonne avec les VRAIS numéros de ligne Excel (1-basés)
        df_raw = df_raw.with_columns(
            pl.Series("original_excel_row", original_row_numbers)
        )

        # 1. Scan des îlots
        ilots = identifier_ilots_donnees(df_raw.drop("original_excel_row"))
        
        tables_found = []
        for i, coord in enumerate(ilots):
            # 2. Extraction chirurgicale
            df_table = df_raw.slice(coord["start_row"], coord["end_row"] - coord["start_row"])
            # Sélectionner les colonnes de données + la colonne de numéros de ligne originaux
            cols_to_keep = ["original_excel_row"] + [
                df_raw.columns[j] for j in range(coord["start_col"], coord["end_col"])
            ]
            df_table = df_table.select(cols_to_keep)

            # 3. Promotion de la première ligne en header
            # On nettoie les lignes totalement vides à l'intérieur du bloc
            df_table = df_table.filter(~pl.all_horizontal(pl.all().exclude("original_excel_row").is_null()))
            
            if not df_table.is_empty():
                # On récupère les headers potentiels
                header_row = df_table.row(0)[1:] # On ignore l'index pour le header
                new_columns = ["original_excel_row"] + [str(h) if h else f"col_{j}" for j, h in enumerate(header_row)]
                
                df_final = df_table.slice(1) # On enlève la ligne de header des données
                df_final.columns = new_columns
                
                title = f"Bloc_{coord['start_row']+1}_to_{coord['end_row']}"
                tables_found.append((title, df_final))

        return tables_found

    except Exception as e:
        print(f"❌ Erreur Segmentation : {e}")
        return []

# ---------------------------------------------------------------------------
# Stockage Parquet (remplace le stockage en mémoire)
# ---------------------------------------------------------------------------

def save_tables_to_parquet(
    tables: List[Tuple[Optional[str], pl.DataFrame]],
    session_id: str
) -> List[Dict]:
    """
    Sauvegarde chaque DataFrame en Parquet dans un répertoire temporaire.

    Retourne une liste de métadonnées :
    [{"title": str, "parquet_path": str, "columns": [...], "n_rows": int}, ...]
    """
    session_dir = PARQUET_TEMP_DIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    metadata_list = []
    for idx, (title, df) in enumerate(tables):
        safe_title = f"table_{idx}" if not title else title.replace("/", "_").replace(" ", "_")
        parquet_path = session_dir / f"{safe_title}.parquet"
        df.write_parquet(str(parquet_path))

        metadata_list.append({
            "title": title or f"Table {idx}",
            "parquet_path": str(parquet_path),
            "columns": df.columns,
            "dtypes": {col: str(df[col].dtype) for col in df.columns},
            "n_rows": len(df),
            "n_cols": len(df.columns),
        })
        print(f"💾 Parquet sauvegardé : {parquet_path} ({len(df)} lignes × {len(df.columns)} cols)")

    return metadata_list


def load_table_from_parquet(parquet_path: str) -> pl.DataFrame:
    """Charge un DataFrame depuis son fichier Parquet."""
    return pl.read_parquet(parquet_path)


def cleanup_session_parquet(session_id: str):
    """Supprime les fichiers Parquet d'une session terminée."""
    session_dir = PARQUET_TEMP_DIR / session_id
    if session_dir.exists():
        for f in session_dir.glob("*.parquet"):
            f.unlink()
        session_dir.rmdir()
        print(f"🧹 Parquet nettoyé pour la session {session_id}")


# ---------------------------------------------------------------------------
# Indexation ChromaDB — MÉTADONNÉES DE COLONNES UNIQUEMENT
# ---------------------------------------------------------------------------

def _build_column_description(meta: Dict) -> str:
    """
    Construit une description ULTRA-LÉGÈRE des colonnes pour l'embedding.
    ✅ UNIQUEMENT le schéma (noms + types) - PAS les données
    ✅ Pas de "sample" pour rester sous le context length
    """
    lines = [f"Tableau: {meta['title']}"]
    lines.append(f"Taille: {meta['n_rows']} lignes, {meta['n_cols']} colonnes")
    lines.append("Colonnes: " + ", ".join([f"{col} ({dtype})" for col, dtype in meta["dtypes"].items()]))
    
    return "\n".join(lines)


def create_vector_store(
    metadata_list: List[Dict],
    session_id: str
) -> Chroma:
    """
    Crée un ChromaDB éphémère indexant UNIQUEMENT les métadonnées de colonnes.

    ✅ Beaucoup plus léger qu'indexer toutes les lignes
    ✅ Le LLM utilise ces métadonnées pour générer du code Polars ciblé
    ✅ Le code généré lit ensuite le Parquet pour produire le résultat réel
    """
    documents = []

    for meta in metadata_list:
        description = _build_column_description(meta)
        doc = Document(
            page_content=description,
            metadata={
                "title": meta["title"],
                "parquet_path": meta["parquet_path"],
                "columns": ", ".join(meta["columns"]),
                "n_rows": meta["n_rows"],
                "session": session_id,
            }
        )
        documents.append(doc)

    # Configuration embedding Ollama - OPTIMISÉ POUR MÉTADONNÉES LÉGÈRES
    embeddings = OllamaEmbeddings(
        model="nomic-embed-text",  # Meilleur pour les textes courts/métadonnées
        base_url=URL_OLLAMA,
    )

    # Test de connectivité Ollama
    try:
        print("⏳ Test Ollama...")
        test_vector = embeddings.embed_query("test connexion")
        print(f"✅ Ollama OK (vecteur taille {len(test_vector)})")
    except Exception as e:
        raise RuntimeError(f"Ollama inaccessible : {e}")

    # Nouveau client éphémère = mémoire précédente effacée
    client = chromadb.EphemeralClient()
    collection_name = f"schema_only_{session_id}"

    try:
        client.delete_collection(name=collection_name)
        print(f"🧹 Ancienne collection '{collection_name}' supprimée.")
    except Exception:
        print(f"✨ Nouvelle collection '{collection_name}'.")

    try:
        vectorstore = Chroma.from_documents(
            documents=documents,
            embedding=embeddings,
            client=client,
            collection_name=collection_name,
        )
        print(f"✅ ChromaDB prêt : {len(documents)} schémas de tableaux indexés (colonnes uniquement)")
    except Exception as e:
        raise RuntimeError(f"ChromaDB : insertion échouée — {e}")

    return vectorstore


# ---------------------------------------------------------------------------
# Point d'entrée principal : pipeline complet
# ---------------------------------------------------------------------------

def process_excel_file(
    uploaded_file,
    sheet_name: str,
    session_id: Optional[str] = None
) -> Tuple[Chroma, List[Dict]]:
    """
    Pipeline complet :
      1. Lecture    → Polars + Calamine (ou openpyxl fallback)
      2. Stockage   → Parquet temporaire
      3. Indexation → ChromaDB sur métadonnées de colonnes

    Retourne (vectorstore, metadata_list).
    Le metadata_list contient les chemins Parquet pour l'exécution du code généré.
    """
    if session_id is None:
        session_id = str(uuid.uuid4())[:8]

    # 1. Lecture
    tables = find_tables_in_sheet(uploaded_file, sheet_name)
    if not tables:
        raise ValueError(f"Aucun tableau trouvé dans la feuille '{sheet_name}'.")
    print(f"📊 {len(tables)} tableau(x) détecté(s)")

    # 2. Stockage Parquet
    metadata_list = save_tables_to_parquet(tables, session_id)

    # 3. Indexation ChromaDB (métadonnées seulement)
    vectorstore = create_vector_store(metadata_list, session_id)

    return vectorstore, metadata_list


# ---------------------------------------------------------------------------
# Fonctions de compatibilité avec l'ancienne API (xlsx_parser.py)
# ---------------------------------------------------------------------------

def get_csv_client(session_id: str = "default") -> chromadb.EphemeralClient:
    """
    Obtient ou crée un client Chroma ÉPHÉMÈRE (in-memory) pour cette session.
    """
    if session_id not in _csv_clients:
        _csv_clients[session_id] = chromadb.EphemeralClient()
        print(f"✅ Nouveau client Chroma éphémère créé pour session {session_id}")
    return _csv_clients[session_id]


def get_csv_collection_name(session_id: str = "default") -> str:
    """Génère un nom de collection unique pour cette session."""
    return f"csv_excel_{session_id}"


def prepare_rag_documents(tables_found: List[Tuple[Optional[str], List[List]]]) -> List[Dict]:
    """
    Transforme les listes de listes (tableaux) en format Markdown pour le RAG.
    Format compatibilité : retourne une liste de dicts {titre: pl.DataFrame}
    """
    documents = []
    for title, data in tables_found:
        # 1. Conversion en DataFrame Polars (data[0] = en-tête, data[1:] = valeurs)
        headers = [str(h) if h is not None else f"col_{i}"
                   for i, h in enumerate(data[0])]
        df = pl.DataFrame(
            {h: [row[i] if i < len(row) else None for row in data[1:]]
             for i, h in enumerate(headers)}
        )
        
        # 2. Ajout comme dict {titre: df}
        documents.append({title or f"Table_{len(documents)}": df})
    
    return documents


def create_vector_store_legacy(
    raw_data_list: List[Dict[str, pl.DataFrame]], 
    session_id: str
) -> Chroma:
    """
    Version héritée de create_vector_store pour compatibilité avec main.py.
    Crée un ChromaDB éphémère indexant UNIQUEMENT LES MÉTADONNÉES (colonnes).
    ✅ Les données complètes resteront dans les Parquets temporaires
    """
    documents = []
    
    # 1. Transformation des DataFrames en Documents (métadonnées seulement)
    for item in raw_data_list:
        for title, df in item.items():
            # Créer une description LÉGÈRE (uniquement colonnes, pas de données)
            dtype_map = {col: str(df[col].dtype) for col in df.columns}
            description = f"Tableau: {title}\nTaille: {len(df)} lignes, {len(df.columns)} colonnes\nColonnes: {', '.join([f'{col} ({dtype})' for col, dtype in dtype_map.items()])}"
            
            # Création du Document LangChain (MÉTADONNÉES SEULEMENT)
            doc = Document(
                page_content=description,
                metadata={
                    "title": title,
                    "n_rows": len(df),
                    "n_cols": len(df.columns),
                    "session": session_id
                }
            )
            documents.append(doc)
    
    # 2. Configuration de l'embedding - OPTIMISÉ POUR MÉTADONNÉES
    embeddings = OllamaEmbeddings(
        model="nomic-embed-text",  # Meilleur pour textes courts
        base_url=URL_OLLAMA
    )
    
    # 3. Créer un client éphémère FRAIS pour cette session
    client = chromadb.EphemeralClient()
    
    # Stocker le client pour les futures requêtes
    _csv_clients[session_id] = client
    
    collection_name = get_csv_collection_name(session_id)
    try:
        client.delete_collection(name=collection_name)
        print(f"🧹 Ancienne collection '{collection_name}' supprimée.")
    except Exception:
        print(f"✨ Création d'une nouvelle collection '{collection_name}'.")
    
    print(f"Préparation de {len(documents)} documents (métadonnées colonnes)...")
    
    try:
        print("⏳ Test de communication avec Ollama...")
        test_vector = embeddings.embed_query("test")
        print(f"✅ Ollama OK (Vecteur {len(test_vector)}d)")
    except Exception as e:
        print(f"❌ CRASH OLLAMA : {e}")
        raise RuntimeError("Ollama est inaccessible depuis le backend.")
    
    try:
        print(f"⏳ Création de la collection Chroma '{collection_name}'...")
        vectorstore = Chroma.from_documents(
            documents=documents,
            embedding=embeddings,
            client=client,
            collection_name=collection_name
        )
        print(f"✅ ChromaDB prêt : {len(documents)} schémas indexés (colonnes uniquement) !")
    except Exception as e:
        raise RuntimeError(f"ChromaDB : insertion échouée — {e}")

    return vectorstore