import pandas as pd
import openpyxl
from typing import List, Dict

def find_tables_in_sheet(uploaded_file, sheet_name: str) -> List[Dict]:
    """
    Détecte automatiquement tous les tableaux dans une feuille Excel.
    
    Retourne une liste de dicts avec:
    - "nom": nom du tableau (Tableau A, Tableau B, etc.)
    - "dataframe": le DataFrame parsé
    - "position": (row_start, col_start) pour le contexte
    """
    
    # Charger le workbook pour analyser la structure brute
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[sheet_name]
    
    # --- Étape 1 : Détecter les "blocs de contenu" (zones sans lignes/colonnes vides) ---
    
    # Récupérer toutes les cellules avec du contenu
    occupied_cells = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                occupied_cells.add((cell.row, cell.column))
    
    if not occupied_cells:
        return []
    
    # Trouver les plages rectangulaires (tableaux) en groupant les cellules
    tableaux = []
    processed = set()
    
    for start_row, start_col in sorted(occupied_cells):
        if (start_row, start_col) in processed:
            continue
        
        # Déterminer les limites du tableau à partir de ce point
        # En cherchant le rectangle maximal de cellules contiguës
        max_row, max_col = start_row, start_col
        
        # Élargir vers la droite tant qu'il y a du contenu
        while (start_row, max_col + 1) in occupied_cells:
            max_col += 1
        
        # Élargir vers le bas tant qu'il y a du contenu (mais attention aux trous)
        row = start_row + 1
        while row <= ws.max_row:
            has_content_in_row = any((row, c) in occupied_cells for c in range(start_col, max_col + 1))
            if not has_content_in_row:
                # Une ligne vide : arrêter (fin du tableau)
                break
            max_row = row
            row += 1
        
        # Marquer comme traité
        for r in range(start_row, max_row + 1):
            for c in range(start_col, max_col + 1):
                processed.add((r, c))
        
        tableaux.append({
            "row_start": start_row,
            "col_start": start_col,
            "row_end": max_row,
            "col_end": max_col,
        })
    
    wb.close()
    
    # --- Étape 2 : Extraire chaque tableau comme DataFrame ---
    
    resultats = []
    
    for idx, tableau_info in enumerate(tableaux):
        row_start = tableau_info["row_start"]
        col_start = tableau_info["col_start"]
        row_end = tableau_info["row_end"]
        col_end = tableau_info["col_end"]
        
        # Lire la plage spécifique
        # openpyxl utilise 1-based indexing, pandas lit par row/col numbers
        # pandas.read_excel accepte des plages Excel-like : "A1:E10"
        
        # Convertir en notation Excel
        from openpyxl.utils import get_column_letter
        col_start_letter = get_column_letter(col_start)
        col_end_letter = get_column_letter(col_end)
        cell_range = f"{col_start_letter}{row_start}:{col_end_letter}{row_end}"
        
        # Lire uniquement cette plage
        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            usecols=f"{col_start_letter}:{col_end_letter}",
            skiprows=row_start - 1,  # pandas skiprows est 0-based
            nrows=row_end - row_start + 1,
            header=0,  # première ligne de la plage = header
        )
        
        # --- Nettoyage du DataFrame ---
        
        # Supprimer colonnes totalement vides ou avec nom "Unnamed"
        df = df.loc[:, df.columns.notna()]
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
        df = df.dropna(how="all", axis=1)
        
        # Supprimer lignes totalement vides
        df = df.dropna(how="all", axis=0)
        
        # Nettoyer les noms de colonnes (accents, espaces)
        df.columns = [
            str(c).strip()
                   .replace(" ", "_")
                   .replace("'", "")
                   .replace("é", "e")
                   .replace("è", "e")
                   .replace("ç", "c")
            for c in df.columns
        ]
        
        # Essayer de reconnaître les types numériques
        for col in df.columns:
            # Si la colonne ressemble à des nombres, les convertir
            try:
                df[col] = pd.to_numeric(df[col], errors="ignore")
            except:
                pass
        
        # Ajouter au résultat avec un nom automatique
        tableau_name = f"Tableau {chr(65 + idx)}"  # A, B, C, ...
        
        resultats.append({
            "nom": tableau_name,
            "dataframe": df,
            "position": (row_start, col_start),
            "shape": df.shape,
        })
    
    return resultats


def afficher_et_charger_tableaux(uploaded_file, sheet_name: str):
    """
    Interface Streamlit pour charger et choisir un tableau.
    Retourne le DataFrame sélectionné.
    """
    import streamlit as st
    
    tableaux = find_tables_in_sheet(uploaded_file, sheet_name)
    
    if not tableaux:
        st.error("❌ Aucun tableau détecté dans cette feuille.")
        return None
    
    if len(tableaux) == 1:
        st.success(f"✅ Un tableau détecté : **{tableaux[0]['nom']}** ({tableaux[0]['shape'][0]} lignes × {tableaux[0]['shape'][1]} colonnes)")
        return tableaux[0]["dataframe"]
    
    else:
        st.info(f"📊 {len(tableaux)} tableaux détectés dans cette feuille.")
        
        # Menu pour choisir le tableau
        options = [f"{t['nom']} — {t['shape'][0]} lignes × {t['shape'][1]} colonnes" for t in tableaux]
        choice_idx = st.radio("Sélectionnez le tableau à analyser :", range(len(options)), format_func=lambda i: options[i])
        
        return tableaux[choice_idx]["dataframe"]