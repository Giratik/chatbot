#backend/robuster.py

import pandas as pd
import os
import openpyxl
from typing import List, Dict
from langchain_core.documents import Document
import uuid
 
import chromadb
from langchain_community.vectorstores import Chroma
from langchain_community.embeddings import OllamaEmbeddings
from chromadb.utils import embedding_functions
from openpyxl.utils import get_column_letter

URL_OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
 
# ── Clients éphémères IN-MEMORY (une par session) ──
_csv_clients: Dict[str, chromadb.EphemeralClient] = {}
 
def get_merged_context(sheet, row, col):
    """Vérifie si une cellule est fusionnée et renvoie sa valeur et sa plage."""
    for merged_range in sheet.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and \
           col >= merged_range.min_col and col <= merged_range.max_col:
            val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return val, merged_range
    return None, None
 
 
def extract_table_with_strict_bounds(sheet, start_row, start_col, visited):
    """Extrait un bloc de données en isolant le titre fusionné."""
    title, m_range = get_merged_context(sheet, start_row, start_col)
    
    # Définit où commencent les colonnes/données
    actual_start_row = m_range.max_row + 1 if m_range else start_row
    
    if m_range:
        # Marquer toute la zone du titre comme visitée pour ne pas reboucler dessus
        for r in range(m_range.min_row, m_range.max_row + 1):
            for c in range(m_range.min_col, m_range.max_col + 1):
                visited.add((r, c))
 
    # 1. Déterminer la largeur basée sur la première ligne après le titre
    max_c = start_col
    while sheet.cell(row=actual_start_row, column=max_c).value is not None:
        max_c += 1
    
    if max_c == start_col:
        return None, None, visited
 
    table_data = []
    curr_row = actual_start_row
    
    # 2. Extraction ligne par ligne
    while curr_row <= sheet.max_row:
        row_values = []
        is_row_empty = True
        
        # Si on tombe sur une NOUVELLE fusion au début, c'est sûrement le titre suivant
        test_val, test_merge = get_merged_context(sheet, curr_row, start_col)
        if curr_row > actual_start_row and test_merge:
            break
 
        for c in range(start_col, max_c):
            val = sheet.cell(row=curr_row, column=c).value
            if val is not None:
                is_row_empty = False
            row_values.append(val)
        
        if is_row_empty:  # Arrêt si ligne vide (fin du tableau)
            break
            
        table_data.append(row_values)
        for c in range(start_col, max_c):
            visited.add((curr_row, c))
        curr_row += 1
        
    return table_data, title, visited
 
# ---------------------------------------------------------------------------
# Détection des tableaux dans une feuille Excel
# ---------------------------------------------------------------------------

def detecter_blocs_continus(df):
    """Compte le nombre de blocs de données séparés par des lignes vides."""
    if df.empty:
        return 0
    lignes_vides = df.isnull().all(axis=1)
    groupes = (lignes_vides != lignes_vides.shift()).cumsum()
    return (~lignes_vides).groupby(groupes).any().sum()


def find_tables_in_sheet(uploaded_file, sheet_name):
    # --- 1. LECTURE RAPIDE (PANDAS/CALAMINE) ---
    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
            
        # Utilisation de calamine pour une vitesse extrême
        df_complet = pd.read_excel(uploaded_file, sheet_name=sheet_name, engine="calamine")
        nb_blocs = detecter_blocs_continus(df_complet)
        
        if nb_blocs == 1:
            print("⚡ Un seul tableau détecté. Extraction instantanée !")
            df_propre = df_complet.dropna(how='all', axis=0).dropna(how='all', axis=1)
            df_propre = df_propre.where(pd.notnull(df_propre), None)
            # Ajout des en-têtes comme première ligne
            table_data = [df_propre.columns.tolist()] + df_propre.values.tolist()
            return [(sheet_name, table_data)]
            
    except Exception as e:
        print(f"⚠️ Échec de la lecture rapide, fallback sur Openpyxl : {e}")

    # --- 2. FALLBACK SUR OPENPYXL (Tableaux multiples ou complexes) ---
    print("🔄 Bascule sur Openpyxl (Analyse spatiale)...")
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)
        
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    tables_found = []
    
    if sheet_name:
        sheet = wb[sheet_name]
        visited = set()
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                if (r, c) not in visited and sheet.cell(row=r, column=c).value is not None:
                    data, title, visited = extract_table_with_strict_bounds(sheet, r, c, visited)
                    if data and len(data) > 1:
                        tables_found.append((title, data))
    
    wb.close()
    return tables_found
 
# ---------------------------------------------------------------------------
# Préparation des documents RAG
# ---------------------------------------------------------------------------
 
def prepare_rag_documents(tables_found):
    """
    Transforme les listes de listes (tableaux) en format Markdown pour le RAG.
    Retourne une liste de dicts {titre: DataFrame}
    """
    documents = []
    for title, data in tables_found:
        # 1. Conversion en DataFrame (data[0] = en-tête, data[1:] = valeurs)
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 2. Ajout comme dict {titre: df}
        documents.append({title: df})
    
    return documents
 
 
# ---------------------------------------------------------------------------
# Gestion du client Chroma éphémère par session
# ---------------------------------------------------------------------------
 
def get_csv_client(session_id: str = "default") -> chromadb.EphemeralClient:
    """
    Obtient ou crée un client Chroma ÉPHÉMÈRE (in-memory) pour cette session.
    ⚠️ Chaque appel crée un nouveau client vierge (pas de persistance).
    """
    if session_id not in _csv_clients:
        _csv_clients[session_id] = chromadb.EphemeralClient()
        print(f"✅ Nouveau client Chroma éphémère créé pour session {session_id}")
    return _csv_clients[session_id]
 
 
def get_csv_collection_name(session_id: str = "default") -> str:
    """Génère un nom de collection unique pour cette session."""
    return f"csv_excel_{session_id}"
 
 
# ---------------------------------------------------------------------------
# Création du vector store (remplace le contenu à chaque appel)
# ---------------------------------------------------------------------------
 
def create_vector_store(raw_data_list: List[Dict[str, pd.DataFrame]], session_id: str) -> Chroma:
    """
    Crée un ChromaDB éphémère avec les DataFrames en Markdown.
    
    ✅ À CHAQUE APPEL : recrée une nouvelle base (l'ancienne est discardée)
    ✅ Pas de persistance sur disque
    ✅ Mémoire libérée quand la session se termine
    """
    
    documents = []
    
    # 1. Transformation des DataFrames en Documents
    for item in raw_data_list:
        for title, df in item.items():
            # Conversion en Markdown (plus naturel pour RAG)
            table_md = df.to_markdown(index=False)
            
            # Création du Document LangChain
            doc = Document(
                page_content=f"Titre: {title}\n\n{table_md}",
                metadata={"source": title, "session": session_id}
            )
            documents.append(doc)
 
    # 2. Configuration de l'embedding
    embeddings = OllamaEmbeddings(
        model="qwen3-embedding:0.6b",
        num_ctx=15000,
        base_url=URL_OLLAMA
    )
    #ollama_ef = embedding_functions.OllamaEmbeddingFunction(
    #url=URL_OLLAMA + "/api/embeddings",
    #model_name="qwen3-embedding:0.6b",
    ##num_ctx=12288
#   )
    # 3. Créer un client éphémère FRAIS pour cette session
    #    ⚠️ Si on veut vraiment vider à chaque appel, on peut créer un nouveau client
    #    Mais on réutilise pour garder la session cohérente
    client = chromadb.EphemeralClient()  # ← NOUVEAU CLIENT = mémoire précédente effacée
    
    collection_name = get_csv_collection_name(session_id)  # Nom unique à chaque appel
    try:
        client.delete_collection(name=collection_name)
        print(f"🧹 Ancienne collection '{collection_name}' supprimée pour faire place au nouveau fichier.")
    except Exception:
        # Si la collection n'existe pas encore (1er fichier), Chroma renvoie une erreur qu'on ignore
        print(f"✨ Création d'une nouvelle collection '{collection_name}'.")
        pass
    #
    print(f"Préparation de {len(documents)} documents...")
        
        # --- TEST 1 : OLLAMA SEUL ---
    try:
        print("⏳ Test de communication avec Ollama...")
        test_vector = embeddings.embed_query("Bonjour le monde")
        print(f"✅ Ollama répond parfaitement ! (Vecteur généré de taille {len(test_vector)})")
    except Exception as e:
        print(f"❌ CRASH OLLAMA : Impossible de générer les embeddings. Détail : {e}")
        raise RuntimeError("Ollama est inaccessible depuis le backend FastAPI.")
        
    # --- TEST 2 : CHROMA SEUL ---
    try:
        print(f"⏳ Création de la collection Chroma '{collection_name}'...")
        vectorstore = Chroma.from_documents(
            documents=documents,
            embedding=embeddings,
            client=client,
            collection_name=collection_name
        )
        print("✅ ChromaDB a bien enregistré les documents !")
    except Exception as e:
        print(f"❌ CRASH CHROMA : Problème avec la base de données. Détail : {e}")
        raise RuntimeError("ChromaDB a refusé l'insertion des documents.")
    ## 4. Création du VectorStore
    #vectorstore = Chroma.from_documents(
    #    documents=documents,
    #    embedding=embeddings,
    #    client=client,
    #    collection_name=collection_name
    #)
    
    print(f"✅ VectorStore créé : {len(documents)} documents indexés")
    return vectorstore


