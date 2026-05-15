#!/usr/bin/env python3
"""
Test de vérification que les numéros de ligne originaux sont correctement préservés.
"""

import sys
import openpyxl
from pathlib import Path
import polars as pl

# Ajouter le backend au chemin
sys.path.insert(0, str(Path(__file__).parent / "backend"))

from new_xlsx_parser import _read_excel_with_original_row_numbers, find_tables_in_sheet


def create_test_excel():
    """Crée un fichier Excel de test avec des lignes vides."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Tableau 1
    ws['A1'] = "Tableau 1"
    ws['A2'] = "Nom"
    ws['B2'] = "Âge"
    ws['A3'] = "Alice"
    ws['B3'] = 25
    ws['A4'] = "Bob"
    ws['B4'] = 30
    
    # Ligne vide (ligne 5)
    
    # Tableau 2
    ws['A6'] = "Tableau 2"
    ws['A7'] = "Produit"
    ws['B7'] = "Prix"
    ws['A8'] = "Pomme"
    ws['B8'] = 1.5
    ws['A9'] = "Banane"
    ws['B9'] = 0.8
    
    wb.save("test_excel.xlsx")
    print("✅ Fichier test créé : test_excel.xlsx")
    return "test_excel.xlsx"


def test_original_row_numbers():
    """Teste que les numéros de ligne originaux sont bien préservés."""
    
    # Créer le fichier de test
    file_path = create_test_excel()
    
    try:
        with open(file_path, "rb") as f:
            df_raw, original_row_numbers = _read_excel_with_original_row_numbers(f, "Test")
        
        print("\n📊 DataFrame avec numéros de ligne originaux:")
        print(df_raw)
        print("\n🔢 Numéros de ligne originaux (Excel):")
        print(original_row_numbers)
        
        # Vérifier que les indices correspondent
        print("\n✅ Correspondance ligne par ligne:")
        for i, (row_num, row_data) in enumerate(zip(original_row_numbers, df_raw.iter_rows())):
            non_empty = any(v is not None for v in row_data)
            if non_empty or i < 5:  # Afficher les 5 premières et les lignes non-vides
                print(f"  Index DF {i} ← Ligne Excel {row_num}: {row_data[:3]}...")
        
    finally:
        Path(file_path).unlink()
        print(f"\n🧹 Fichier test supprimé")


def test_segmentation_with_row_numbers():
    """Teste que la segmentation préserve correctement les numéros de ligne."""
    
    file_path = create_test_excel()
    
    try:
        with open(file_path, "rb") as f:
            tables = find_tables_in_sheet(f, "Test")
        
        print("\n📋 Tableaux extraits avec numéros de ligne originaux:")
        for title, df_table in tables:
            print(f"\n  {title}:")
            print(f"    Colonnes: {df_table.columns}")
            print(f"    Données:\n{df_table}")
            
            # Afficher les numéros de ligne originaux
            if "original_excel_row" in df_table.columns:
                original_rows = df_table["original_excel_row"].to_list()
                print(f"    ✅ Numéros de ligne originaux (Excel): {original_rows}")
            else:
                print(f"    ❌ Colonne 'original_excel_row' non trouvée!")
    
    finally:
        Path(file_path).unlink()


if __name__ == "__main__":
    print("=" * 60)
    print("TEST 1: Vérification des numéros de ligne originaux")
    print("=" * 60)
    test_original_row_numbers()
    
    print("\n" + "=" * 60)
    print("TEST 2: Segmentation avec numéros de ligne originaux")
    print("=" * 60)
    test_segmentation_with_row_numbers()
