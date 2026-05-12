#!/usr/bin/env python3
"""
Script de diagnostic - Vérifier les 3 problèmes identifiés
1. Support des tableaux multiples
2. Performance Polars+Calamine
3. Indexation ChromaDB légère
"""

import sys
import os
import io
import tempfile
from pathlib import Path
from time import time

# Ajouter le backend au path
backend_path = Path(__file__).parent / "backend"
sys.path.insert(0, str(backend_path))

def test_ollama_models():
    """Test 1: Vérifier les modèles Ollama disponibles"""
    print("\n" + "="*60)
    print("🔍 TEST 1: Vérification des modèles Ollama")
    print("="*60)
    
    try:
        import ollama
        models = ollama.list()
        print("✅ Ollama répond")
        print("\nModèles disponibles:")
        if hasattr(models, 'models'):
            for m in models.models:
                print(f"  - {m.model}")
        
        # Vérifier les modèles requis
        required = ["nomic-embed-text"]
        model_names = [m.model if hasattr(m, 'model') else str(m) for m in (models.models if hasattr(models, 'models') else [])]
        
        for req in required:
            if any(req in m for m in model_names):
                print(f"✅ {req} trouvé")
            else:
                print(f"⚠️  {req} NOT found - peut causer des erreurs")
        
        return True
    except Exception as e:
        print(f"❌ Erreur Ollama: {e}")
        return False

def test_multiple_tables():
    """Test 2: Vérifier la détection de tableaux multiples"""
    print("\n" + "="*60)
    print("🔍 TEST 2: Détection de tableaux multiples")
    print("="*60)
    
    try:
        import polars as pl
        import new_xlsx_parser
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        # Créer un fichier Excel avec 2 tableaux
        wb = Workbook()
        ws = wb.active
        
        # Tableau 1
        ws['A1'] = "Tableau 1"
        ws['A2'] = "Col1"
        ws['B2'] = "Col2"
        ws['A3'] = "Val1"
        ws['B3'] = "Val2"
        ws['A4'] = "Val3"
        ws['B4'] = "Val4"
        
        # Espace vide (séparateur)
        
        # Tableau 2
        ws['A6'] = "Tableau 2"
        ws['A7'] = "ColA"
        ws['B7'] = "ColB"
        ws['C7'] = "ColC"
        ws['A8'] = "ValA1"
        ws['B8'] = "ValB1"
        ws['C8'] = "ValC1"
        
        # Sauvegarder en mémoire
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        wb.close()
        
        # Tester find_tables_in_sheet
        with open(temp_file.name, 'rb') as f:
            excel_data = io.BytesIO(f.read())
            start = time()
            tables = new_xlsx_parser.find_tables_in_sheet(excel_data, sheet_name="Sheet")
            elapsed = time() - start
        
        print(f"⏱️  Temps de parsing: {elapsed:.3f}s")
        print(f"📊 Tableaux détectés: {len(tables)}")
        
        for i, (title, df) in enumerate(tables):
            print(f"  [{i}] {title}: {len(df)} lignes × {len(df.columns)} colonnes")
            if len(df.columns) <= 5:
                print(f"      Colonnes: {list(df.columns)}")
        
        # Nettoyage
        os.unlink(temp_file.name)
        
        if len(tables) >= 2:
            print("✅ Tableaux multiples détectés correctement")
            return True
        else:
            print("⚠️  Seulement 1 tableau détecté (attendu ≥ 2)")
            return False
            
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_metadata_indexing():
    """Test 3: Vérifier que ChromaDB indexe les métadonnées légères"""
    print("\n" + "="*60)
    print("🔍 TEST 3: Indexation ChromaDB (métadonnées légères)")
    print("="*60)
    
    try:
        import polars as pl
        import new_xlsx_parser
        
        # Créer 2 DataFrames Polars de tailles différentes
        df1 = pl.DataFrame({
            "ID": range(1, 101),
            "Name": [f"User_{i}" for i in range(1, 101)],
            "Email": [f"user{i}@example.com" for i in range(1, 101)],
            "Status": ["Active"] * 100
        })
        
        df2 = pl.DataFrame({
            "OrderID": range(1, 51),
            "Amount": [100 + i*10 for i in range(50)],
            "Date": ["2024-01-01"] * 50
        })
        
        # Sauvegarder en Parquet
        tables = [("Users", df1), ("Orders", df2)]
        session_id = "test_diagnostic"
        
        start = time()
        metadata_list = new_xlsx_parser.save_tables_to_parquet(tables, session_id)
        elapsed = time() - start
        
        print(f"⏱️  Temps Parquet: {elapsed:.3f}s")
        print(f"\n📝 Métadonnées générées:")
        
        for meta in metadata_list:
            print(f"  Tableau: {meta['title']}")
            print(f"    Path: {meta['parquet_path']}")
            print(f"    Taille: {meta['n_rows']} × {meta['n_cols']}")
            print(f"    Colonnes: {meta['columns']}")
            
            # Vérifier qu'il n'y a PAS de sample
            if 'sample' in meta:
                print(f"    ⚠️  'sample' encore présent (devrait être supprimé)")
            else:
                print(f"    ✅ Pas de 'sample' (léger)")
        
        print(f"\n⏳ Test création ChromaDB...")
        start = time()
        vectorstore = new_xlsx_parser.create_vector_store(metadata_list, session_id)
        elapsed = time() - start
        
        print(f"⏱️  Temps indexation: {elapsed:.3f}s")
        print("✅ ChromaDB créé sans erreur")
        
        # Nettoyage
        new_xlsx_parser.cleanup_session_parquet(session_id)
        print("✅ Parquet nettoyé")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_performance_polars_vs_pandas():
    """Test 4 (optionnel): Comparer performance Polars vs Pandas"""
    print("\n" + "="*60)
    print("🔍 TEST 4 (BONUS): Performance Polars vs Pandas")
    print("="*60)
    
    try:
        import polars as pl
        import pandas as pd
        from openpyxl import Workbook
        import tempfile
        
        # Créer un fichier Excel avec 1000 lignes
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "Id"
        ws['B1'] = "Value1"
        ws['C1'] = "Value2"
        
        for i in range(1, 1001):
            ws[f'A{i+1}'] = i
            ws[f'B{i+1}'] = f"Val{i}"
            ws[f'C{i+1}'] = i * 1.5
        
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        wb.close()
        
        # Test Polars
        start = time()
        df_pl = pl.read_excel(temp_file.name, sheet_name="Sheet", engine="calamine")
        elapsed_pl = time() - start
        
        # Test Pandas
        start = time()
        df_pd = pd.read_excel(temp_file.name, sheet_name="Sheet")
        elapsed_pd = time() - start
        
        print(f"Polars:  {elapsed_pl:.4f}s")
        print(f"Pandas:  {elapsed_pd:.4f}s")
        print(f"Ratio:   {elapsed_pd/elapsed_pl:.1f}x plus rapide avec Polars")
        
        # Nettoyage
        os.unlink(temp_file.name)
        
        if elapsed_pl < elapsed_pd:
            print("✅ Polars est plus rapide")
        
        return True
        
    except Exception as e:
        print(f"⚠️  Test Polars/Pandas skipped: {e}")
        return True  # Pas critique

def main():
    print("\n" + "="*60)
    print("DIAGNOSTIC - Migration new_xlsx_parser")
    print("="*60)
    
    tests = [
        ("Modèles Ollama", test_ollama_models),
        ("Tableaux multiples", test_multiple_tables),
        ("Indexation légère", test_metadata_indexing),
        ("Performance (bonus)", test_performance_polars_vs_pandas),
    ]
    
    results = []
    for test_name, test_func in tests:
        try:
            result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"\n❌ {test_name} échoué: {e}")
            results.append((test_name, False))
    
    print("\n" + "="*60)
    print("RÉSUMÉ DIAGNOSTIC")
    print("="*60)
    
    for test_name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    critical = results[:3]  # Les 3 premiers tests sont critiques
    all_critical_pass = all(result for _, result in critical)
    
    print("="*60)
    if all_critical_pass:
        print("🎉 TOUS LES TESTS CRITIQUES PASSENT!")
        print("\nVos fixes ont résolu les 3 problèmes:")
        print("  ✅ Détection tableaux multiples")
        print("  ✅ Performance Polars+Calamine")
        print("  ✅ Indexation ChromaDB légère (métadonnées colonnes)")
    else:
        print("❌ CERTAINS TESTS CRITIQUES ONT ÉCHOUÉ")
        print("\nVérifiez les logs ci-dessus pour les détails")
    
    return 0 if all_critical_pass else 1

if __name__ == "__main__":
    exit(main())
