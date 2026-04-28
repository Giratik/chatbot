import streamlit as st
import pandas as pd
import chromadb
from langchain_community.vectorstores import Chroma
from langchain_community.embeddings import OllamaEmbeddings
from langchain_core.documents import Document
from openpyxl.utils import get_column_letter
import openpyxl

from streamlit_option_menu import option_menu

import uuid

import os
import httpx
from ollama import Client
from typing import List, Dict

# ---------------------------------------------------------------------------
# Configuration Ollama
# ---------------------------------------------------------------------------

URL_OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
client = Client(
    host=URL_OLLAMA,
    timeout=httpx.Timeout(
        connect=5.0,
        read=600.0,
        write=10.0,
        pool=5.0
    )
)



def get_merged_context(sheet, row, col):
    """Vérifie si une cellule est fusionnée et renvoie sa valeur et sa plage."""
    for merged_range in sheet.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and \
           col >= merged_range.min_col and col <= merged_range.max_col:
            val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return val, merged_range
    return None, None


def extract_table_with_strict_bounds(sheet, start_row, start_col, visited):
    """Extrait un bloc de données en isolant le titre fusionné."""
    title, m_range = get_merged_context(sheet, start_row, start_col)
    
    # Définit où commencent les colonnes/données
    actual_start_row = m_range.max_row + 1 if m_range else start_row
    
    if m_range:
        # Marquer toute la zone du titre comme visitée pour ne pas reboucler dessus
        for r in range(m_range.min_row, m_range.max_row + 1):
            for c in range(m_range.min_col, m_range.max_col + 1):
                visited.add((r, c))

    # 1. Déterminer la largeur basée sur la première ligne après le titre
    max_c = start_col
    while sheet.cell(row=actual_start_row, column=max_c).value is not None:
        max_c += 1
    
    if max_c == start_col:
        return None, None, visited

    table_data = []
    curr_row = actual_start_row
    
    # 2. Extraction ligne par ligne
    while curr_row <= sheet.max_row:
        row_values = []
        is_row_empty = True
        
        # Si on tombe sur une NOUVELLE fusion au début, c'est sûrement le titre suivant
        test_val, test_merge = get_merged_context(sheet, curr_row, start_col)
        if curr_row > actual_start_row and test_merge:
            break

        for c in range(start_col, max_c):
            val = sheet.cell(row=curr_row, column=c).value
            if val is not None:
                is_row_empty = False
            row_values.append(val)
        
        if is_row_empty: # Arrêt si ligne vide (fin du tableau)
            break
            
        table_data.append(row_values)
        for c in range(start_col, max_c):
            visited.add((curr_row, c))
        curr_row += 1
        
    return table_data, title, visited

# ---------------------------------------------------------------------------
# Détection des tableaux dans une feuille Excel
# ---------------------------------------------------------------------------

def find_tables_in_sheet(uploaded_file, sheet_name):
    """
    Détecte automatiquement tous les tableaux dans une feuille Excel.
    Retourne une liste de tuples (titre, donnees_brutes).
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    tables_found = []
    
    if sheet_name:
        sheet = wb[sheet_name]
        st.info(f"Analyse de la feuille : **{sheet_name}** ({sheet.max_row} lignes, {sheet.max_column} colonnes)")

        # RETRAIT DU ST.BUTTON ICI
        visited = set()

        # Parcours de la grille de la feuille sélectionnée
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                if (r, c) not in visited and sheet.cell(row=r, column=c).value is not None:
                    data, title, visited = extract_table_with_strict_bounds(sheet, r, c, visited)
                    if data and len(data) > 1:
                        tables_found.append((title, data))
        
    return tables_found


# ---------------------------------------------------------------------------
# Préparation des documents RAG
# ---------------------------------------------------------------------------

def prepare_rag_documents(tables_found):
    """
    Transforme les listes de listes (tableaux) en format Markdown pour le RAG.
    """
    documents = []
    # CORRECTION ICI : on déballe le tuple (title, data)
    for i, (title, data) in enumerate(tables_found):
        # 1. Conversion en DataFrame (data[0] est bien la ligne d'en-tête, data[1:] les valeurs)
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 2. Conversion en Markdown
        #table_md = df.to_markdown(index=False)
        
        # 3. Création d'un document avec métadonnées
        # J'en profite pour inclure le vrai titre du tableau dans le page_content et les métadonnées
        titre_affichage = title
        doc = {title : df}
        documents.append(doc)
    return documents


# ---------------------------------------------------------------------------
# Création du vector store ChromaDB (éphémère, en RAM)
# ---------------------------------------------------------------------------

def create_vector_store(raw_data_list: List[Dict[str, pd.DataFrame]]) -> Chroma:
    """
    Transforme les DataFrames en texte (Markdown) puis les indexe dans 
    un ChromaDB éphémère (100 % en mémoire) via Ollama.
    """
    documents = []
    
    # 1. Traitement des dictionnaires {titre: DataFrame}
    for item in raw_data_list:
        for title, df in item.items():
            # Conversion du DataFrame en texte (Markdown) pour l'embedding
            table_md = df.to_markdown(index=False)
            
            # Création du Document attendu par LangChain/Chroma
            doc = Document(
                page_content=f"Titre du tableau : {title}\n\n{table_md}",
                metadata={"source": title, "type": "table"}
            )
            documents.append(doc)

    # 2. Configuration de l'embedding
    embeddings = OllamaEmbeddings(
        model="qwen3-embedding:0.6b",
        num_ctx=12288
    )

    ephemeral_client = chromadb.EphemeralClient()

    # SOLUTION ICI : Générer un nom unique pour forcer une collection vierge
    nom_unique = f"rag_excel_{uuid.uuid4().hex}"

    vectorstore = Chroma.from_documents(
        documents=documents,
        embedding=embeddings,
        client=ephemeral_client,
        collection_name=nom_unique  # On utilise le nom unique !
    )

    return vectorstore  


# ---------------------------------------------------------------------------
# Inférence Ollama
# ---------------------------------------------------------------------------

def inferring_ollama(messages, model, temperature=0.4, stream=False, context_size=32768):
    """Envoie les messages à Ollama et retourne la réponse texte."""
    response = client.chat(
        model=model,
        messages=messages,
        options={
            "temperature": temperature,
            "num_ctx": context_size,
        },
        stream=stream
    )

    if not stream:
        return response["message"]["content"]


# ---------------------------------------------------------------------------
# Interface Streamlit
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Excel RAG", layout="wide")
st.title("📂 Extracteur de Tableaux Excel — RAG")

# Initialisation de la session
if "vector_index" not in st.session_state:
    st.session_state.vector_index = None
if "resultats" not in st.session_state:
    st.session_state.resultats = None

# 1. Chargement du fichier
uploaded_file = st.file_uploader("Charger un fichier Excel (.xlsx)", type="xlsx")

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    sheet_name = st.selectbox("Choisissez l'onglet :", xls.sheet_names)
    #choice = st.radio("blah", xls.sheet_names, horizontal=True, label_visibility="collapsed")
    #st.caption("choix onglet :", choice)
    with st.chat_message("assistant"):
        with st.container():
            col1, col2 = st.columns(2)
            with col1:
                st.button('Predict sentiment')
            with col2:
                st.button('Predict Price')
        st.text("Quelle page à analyser ?")
        #st.text("Quelle page à analyser ?")
        st.radio("Quelle page à analyser ?", xls.sheet_names, horizontal=True, label_visibility="collapsed")
    

    # 2. Détection des tableaux
    if st.button("🔍 Détecter les tableaux"):
        with st.spinner("Analyse de la feuille en cours..."):
            st.session_state.resultats = find_tables_in_sheet(uploaded_file, sheet_name)
            st.session_state.vector_index = None  # reset si nouvelle détection

        if st.session_state.resultats:
            st.success(f"{len(st.session_state.resultats)} tableau(x) détecté(s).")
        else:
            st.warning("Aucun tableau structuré trouvé sur cette feuille.")

        # 3. Affichage des tableaux détectés
        if st.session_state.resultats:
            st.success(f"{len(st.session_state.resultats)} tableau(x) détecté(s).")
            for i, (title, data) in enumerate(st.session_state.resultats):
                with st.expander(f"Tableau n°{i+1} : {title if title else 'Sans titre'}", expanded=True):
                    # Création du DataFrame (Ligne 0 = Headers)
                    df = pd.DataFrame(data[1:], columns=data[0])
                    st.dataframe(df, use_container_width=True)
                    
                    # Option export CSV par tableau
                    csv = df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="Télécharger en CSV",
                        data=csv,
                        file_name=f"{sheet_name}_table_{i+1}.csv",
                        mime="text/csv",
                        key=f"btn_{sheet_name}_{i}"
                    )
        else:
            st.warning("Aucun tableau structuré n'a été trouvé sur cette feuille.")

    # 4. Indexation RAG
    if st.button("🧠 Indexer pour le RAG"):
        docs = prepare_rag_documents(st.session_state.resultats)
        with st.spinner("Génération des embeddings et indexation ChromaDB..."):
            st.session_state.vector_index = create_vector_store(docs)
        st.success(f"{len(docs)} document(s) indexé(s) — base vectorielle prête !")

    # 5. Interface de question / réponse
    if st.session_state.vector_index:
        st.divider()
        st.subheader("💬 Posez une question sur vos tableaux")
        query = st.text_input("Votre question :")

        if query:
            docs_trouves = st.session_state.vector_index.similarity_search(query, k=10)
            contexte = "\n\n".join(doc.page_content for doc in docs_trouves)

            prompt = f"""Tu es un assistant expert en analyse de données tabulaires.
Utilise uniquement le contexte ci-dessous (tableaux au format Markdown) pour répondre.
Si l'information est absente, réponds : "Je ne trouve pas l'information dans les tableaux."

Contexte :
{contexte}

Question : {query}

Réponse :"""

            messages = [{"role": "user", "content": prompt}]

            with st.spinner("Génération de la réponse..."):
                response = inferring_ollama(messages, model="ministral-3:14b", context_size=12288)

            st.markdown("### 🤖 Réponse")
            st.write(response)

            with st.expander("📄 Voir les sources utilisées"):
                for doc in docs_trouves:
                    st.caption(f"Source : {doc.metadata.get('source', '—')}")
                    st.code(doc.page_content, language="markdown")