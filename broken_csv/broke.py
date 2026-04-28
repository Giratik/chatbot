import streamlit as st
import openpyxl
import pandas as pd

def get_merged_context(sheet, row, col):
    for merged_range in sheet.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and \
           col >= merged_range.min_col and col <= merged_range.max_col:
            val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return val, merged_range
    return None, None

def extract_table_with_strict_bounds(sheet, start_row, start_col, visited):
    title, m_range = get_merged_context(sheet, start_row, start_col)
    
    # Initialisation du point de départ des données
    actual_start_row = m_range.max_row + 1 if m_range else start_row
    
    # On marque le titre comme visité
    if m_range:
        for r in range(m_range.min_row, m_range.max_row + 1):
            for c in range(m_range.min_col, m_range.max_col + 1):
                visited.add((r, c))

    # 1. Déterminer la largeur exacte du header (ligne juste après le titre)
    # On s'arrête dès qu'on trouve une cellule vide à droite
    max_c = start_col
    while sheet.cell(row=actual_start_row, column=max_c).value is not None:
        max_c += 1
    
    # Sécurité : si la ligne sous le titre est vide, ce n'est pas un tableau
    if max_c == start_col:
        return None, None, visited

    table_data = []
    curr_row = actual_start_row
    
    # 2. Extraction avec détection de fin de tableau
    # On s'arrête si :
    # - La ligne est totalement vide sur la largeur définie
    # - OU on rencontre une nouvelle cellule fusionnée (qui serait un nouveau titre)
    while curr_row <= sheet.max_row:
        row_values = []
        is_row_empty = True
        
        # Avant d'extraire la ligne, on vérifie si elle commence par une nouvelle fusion
        # (Indice qu'on a changé de section/page)
        test_val, test_merge = get_merged_context(sheet, curr_row, start_col)
        if curr_row > actual_start_row and test_merge:
            break

        for c in range(start_col, max_c):
            val = sheet.cell(row=curr_row, column=c).value
            if val is not None:
                is_row_empty = False
            row_values.append(val)
        
        if is_row_empty:
            break
            
        table_data.append(row_values)
        for c in range(start_col, max_c):
            visited.add((curr_row, c))
        curr_row += 1
        
    return table_data, title, visited

# --- UI Streamlit ---
st.set_page_config(layout="wide")
st.title("🔍 Extracteur Strict par Bloc")

uploaded_file = st.file_uploader("Fichier Excel", type="xlsx")

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet = wb.active # Ou selectbox pour choisir
    
    visited = set()
    tables_found = []

    # Parcours systématique
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if (r, c) not in visited and sheet.cell(row=r, column=c).value is not None:
                data, title, visited = extract_table_with_strict_bounds(sheet, r, c, visited)
                if data and len(data) > 1:
                    tables_found.append((title, data))

    # Affichage des résultats
    if tables_found:
        for i, (title, data) in enumerate(tables_found):
            with st.container():
                st.subheader(f"Tableau {i+1} : {title if title else 'Sans titre'}")
                df = pd.DataFrame(data[1:], columns=data[0])
                st.dataframe(df, use_container_width=True)
                st.write(f"Nombre de lignes : {len(df)}")
                st.divider()
    else:
        st.info("Aucun tableau structuré trouvé.")