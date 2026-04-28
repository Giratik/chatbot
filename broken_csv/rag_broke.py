import streamlit as st
import pandas as pd
from langchain_community.vectorstores import FAISS
from langchain_core.documents import Document
from langchain_community.embeddings import OllamaEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
import openpyxl

import os
import httpx
from ollama import Client

URL_OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
client = Client(
    host=URL_OLLAMA,
    timeout=httpx.Timeout(
        connect=5.0,    # Connexion au serveur
        read=600.0,     # Attente de la réponse (le plus important)
        write=10.0,     # Envoi du prompt
        pool=5.0        # Attente d'une connexion disponible
    )
)


def inferring_ollama(messages, model, temperature=0.4, stream=False, stats_dict=None, context_size=32768, **kwargs):
    # Appel à l'API avec le paramètre stream
    #with llm_latency.time():
    response = client.chat(
        model=model,
        messages=messages,
        options={
            "temperature": temperature,
            "num_ctx": context_size,
            #"num_predict": max_tokens  # Limite les tokens générés
        },
        stream=stream 
    )
    #if hasattr(response, 'eval_count'):
    #            llm_tokens_generated.inc(response.eval_count)

    if stream==False:
        return response['message']['content']

    

def prepare_rag_documents(tables_found):
    """
    Transforme les listes de listes (tableaux) en format Markdown pour le RAG.
    """
    documents = []
    # CORRECTION ICI : on déballe le tuple (title, data)
    for i, (title, data) in enumerate(tables_found):
        # 1. Conversion en DataFrame (data[0] est bien la ligne d'en-tête, data[1:] les valeurs)
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 2. Conversion en Markdown
        table_md = df.to_markdown(index=False)
        
        # 3. Création d'un document avec métadonnées
        # J'en profite pour inclure le vrai titre du tableau dans le page_content et les métadonnées
        titre_affichage = title if title else f"Tableau n°{i+1}"
        doc = Document(
            page_content=f"{titre_affichage}\n{table_md}",
            metadata={"source": f"Tableau_{i+1}", "type": "table", "title": titre_affichage}
        )
        documents.append(doc)
    return documents

def create_local_vector_store(documents):
    """Crée un index FAISS en utilisant Ollama en local."""
    
    # Configuration de l'embedding local
    # Assurez-vous que l'application Ollama tourne en arrière-plan sur votre machine
    embeddings = OllamaEmbeddings(
        model="qwen3-embedding:0.6b",
        num_ctx=12288  # Force Ollama à utiliser le maximum de tokens possible
    )
    
    # Création du store FAISS
    vectorstore = FAISS.from_documents(documents, embeddings)
    return vectorstore



def get_merged_context(sheet, row, col):
    """Vérifie si une cellule est fusionnée et renvoie sa valeur et sa plage."""
    for merged_range in sheet.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and \
           col >= merged_range.min_col and col <= merged_range.max_col:
            val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return val, merged_range
    return None, None

def extract_table_with_strict_bounds(sheet, start_row, start_col, visited):
    """Extrait un bloc de données en isolant le titre fusionné."""
    title, m_range = get_merged_context(sheet, start_row, start_col)
    
    # Définit où commencent les colonnes/données
    actual_start_row = m_range.max_row + 1 if m_range else start_row
    
    if m_range:
        # Marquer toute la zone du titre comme visitée pour ne pas reboucler dessus
        for r in range(m_range.min_row, m_range.max_row + 1):
            for c in range(m_range.min_col, m_range.max_col + 1):
                visited.add((r, c))

    # 1. Déterminer la largeur basée sur la première ligne après le titre
    max_c = start_col
    while sheet.cell(row=actual_start_row, column=max_c).value is not None:
        max_c += 1
    
    if max_c == start_col:
        return None, None, visited

    table_data = []
    curr_row = actual_start_row
    
    # 2. Extraction ligne par ligne
    while curr_row <= sheet.max_row:
        row_values = []
        is_row_empty = True
        
        # Si on tombe sur une NOUVELLE fusion au début, c'est sûrement le titre suivant
        test_val, test_merge = get_merged_context(sheet, curr_row, start_col)
        if curr_row > actual_start_row and test_merge:
            break

        for c in range(start_col, max_c):
            val = sheet.cell(row=curr_row, column=c).value
            if val is not None:
                is_row_empty = False
            row_values.append(val)
        
        if is_row_empty: # Arrêt si ligne vide (fin du tableau)
            break
            
        table_data.append(row_values)
        for c in range(start_col, max_c):
            visited.add((curr_row, c))
        curr_row += 1
        
    return table_data, title, visited

# --- Interface Streamlit ---
st.set_page_config(page_title="Multi-Sheet Excel Extractor", layout="wide")
st.title("📂 Extracteur de Tableaux Multi-Feuilles")

# 1. Initialiser le stockage dans la session
if "vector_index" not in st.session_state:
    st.session_state.vector_index = None
    
if "tables_found" not in st.session_state:
    st.session_state.tables_found = None


uploaded_file = st.file_uploader("Charger un fichier Excel (.xlsx)", type="xlsx")

if uploaded_file:
    # On charge le classeur une seule fois (data_only=True pour avoir les valeurs et non les formules)
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    # Sélection de la feuille par l'utilisateur
    sheet_name = st.selectbox("Sélectionnez la feuille à analyser :", wb.sheetnames)
    
    if sheet_name:
        sheet = wb[sheet_name]
        st.info(f"Analyse de la feuille : **{sheet_name}** ({sheet.max_row} lignes, {sheet.max_column} colonnes)")

        if st.button("Lancer la détection sur cette feuille"):
            visited = set()
            st.session_state.tables_found = []

            # Parcours de la grille de la feuille sélectionnée
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    if (r, c) not in visited and sheet.cell(row=r, column=c).value is not None:
                        data, title, visited = extract_table_with_strict_bounds(sheet, r, c, visited)
                        if data and len(data) > 1:
                            st.session_state.tables_found.append((title, data))

            # Affichage des résultats
            if st.session_state.tables_found:
                st.success(f"{len(st.session_state.tables_found)} tableau(x) détecté(s).")
                for i, (title, data) in enumerate(st.session_state.tables_found):
                    with st.expander(f"Tableau n°{i+1} : {title if title else 'Sans titre'}", expanded=True):
                        # Création du DataFrame (Ligne 0 = Headers)
                        df = pd.DataFrame(data[1:], columns=data[0])
                        st.dataframe(df, use_container_width=True)
                        
                        # Option export CSV par tableau
                        csv = df.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="Télécharger en CSV",
                            data=csv,
                            file_name=f"{sheet_name}_table_{i+1}.csv",
                            mime="text/csv",
                            key=f"btn_{sheet_name}_{i}"
                        )
            else:
                st.warning("Aucun tableau structuré n'a été trouvé sur cette feuille.")
        
        if st.button("Lancer la détection et préparer le RAG"):
        # (Extraction des tables...)
            #tables_found = [data1, data2] # Données extraites de vos fonctions précédentes
            
            if st.session_state.tables_found:
                # Transformation
                docs = prepare_rag_documents(st.session_state.tables_found)
                
                # Indexation (Stockage temporaire dans st.session_state)
                with st.spinner("Indexation des tableaux en cours..."):
                    st.session_state.vector_index = create_local_vector_store(docs)
                    st.success("Base de données vectorielle prête pour la session !")

    if st.session_state.vector_index:
        st.divider()
        query = st.text_input("Posez une question sur vos tableaux Excel :")
        
        if query:
            # 1. On effectue la recherche sémantique manuellement
            docs_trouves = st.session_state.vector_index.similarity_search(query, k=2)
            
            # 2. On extrait le texte (Markdown) de ces documents et on les assemble
            contexte_extrait = "\n\n".join([doc.page_content for doc in docs_trouves])
            
            # 3. On crée le prompt avec le VRAI contexte
            prompt = f"""
            Tu es un assistant expert en analyse de données.
            Utilise uniquement le contexte suivant (qui est un tableau Markdown) pour répondre à la question.
            Si la réponse ne s'y trouve pas, dis simplement "Je ne trouve pas l'information dans le tableau".
            
            Contexte :
            {contexte_extrait}
            
            Question : {query}
            
            Réponse :
            """
            
            messages_pour_ollama = [{"role": "system", "content": prompt}]
            
            with st.spinner("Analyse du tableau et génération de la réponse..."):
                # 4. On lance l'inférence !
                response = inferring_ollama(messages_pour_ollama, "ministral-3:14b", context_size=12288)
                
                # Affichage de la réponse finale de l'IA
                st.markdown("### 🤖 Réponse :")
                st.write(response)
                
                # Optionnel : Afficher les sources (le fameux tableau brut)
                with st.expander("Voir le tableau source utilisé"):
                    for doc in docs_trouves:
                        st.code(doc.page_content, language="markdown")
                #        st.code(doc.page_content, language="markdown")